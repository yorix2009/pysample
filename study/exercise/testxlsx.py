# coding: UTF-8
#安装 pip3 install openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime


def write_xls():
    wb = Workbook()
    # grab the active worksheet
    ws = wb.active
    # 直接单元格赋值
    ws['A1'] = 42
    # 在下一行插入数据
    ws.append([1, 2, 3])
    # Python types will automatically be converted
    ws['A2'] = datetime.datetime.now()
    # Save the file
    wb.save("sample.xlsx")


def read_xls():
    wb = load_workbook('sample.xlsx')
    ws = wb.active
    names = wb.sheetnames
    print(names)
    print(ws['A1'].value)
    for cell in ws['A2:F20']:
        for x in cell:
            print(x.column,x.row,x.coordinate,x.value)


if __name__ == '__main__':
    read_xls()
